import openpyxl
for path,sh in [("produccion-10-07/CAJA 10-07.xlsx",None)]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    print("=== ",path, ws.title, ws.max_row,"x",ws.max_column)
    for r in range(1, ws.max_row+1):
        row=[]
        for c in range(1, ws.max_column+1):
            v=ws.cell(r,c).value
            if v is not None: row.append(f"{c}:{v!r}")
        if row: print(f"R{r}", " | ".join(row))
