import openpyxl

# ---------- DAY 09 MANUAL CAJA ----------
wb=openpyxl.load_workbook("produccion-09-07/Caja_HyC_09-07-2026.xlsx",data_only=True)
ws=wb.active
section="VENTAS"
ventas_ef=ventas_cr=ventas_dy=0.0
otros=[]; salidas=[]; mixed=[]
for r in range(4,ws.max_row+1):
    a=ws.cell(r,1).value; item=ws.cell(r,2).value; cli=ws.cell(r,3).value
    ef=ws.cell(r,4).value or 0; cr=ws.cell(r,5).value or 0; dy=ws.cell(r,6).value or 0
    if a in ('BOLETAS Y FACTURAS','OTROS INGRESOS','SALIDAS'):
        section=a; continue
    if item=='TOTAL' or a=='CAJA': continue
    if not any([item,cli]) and not any([ef,cr,dy]): continue
    if section in ('VENTAS','BOLETAS Y FACTURAS'):
        ventas_ef+=ef; ventas_cr+=cr; ventas_dy+=dy
        nz=[x for x in [('EF',ef),('CR',cr),('DY',dy)] if x[1]]
        if len(nz)>1: mixed.append((item,cli,ef,cr,dy))
    elif section=='OTROS INGRESOS':
        otros.append((cli,ef,cr,dy))
    elif section=='SALIDAS':
        salidas.append((cli,ef))  # negatives in col4
print("=== DAY09 MANUAL CAJA ===")
print(f"ventas EFECTIVO={ventas_ef}  CREDITO={ventas_cr}  DEP/YAPE={ventas_dy}")
print("OTROS INGRESOS:",otros)
print("SALIDAS:",salidas)
sal_sum=sum(s[1] for s in salidas)
otros_ef=sum(o[1] for o in otros)
print("suma salidas(col4)=",sal_sum," otros_ingresos_ef=",otros_ef)
print("MIXED:",mixed)
inicio=11038.79
fin_calc=inicio+ventas_ef+otros_ef+sal_sum   # salidas already negative
print("fin_calc=",round(fin_calc,2)," oficial=13591.69  diff=",round(fin_calc-13591.69,2))
print()

# ---------- SYSTEM EXPORTS ----------
def analyze_system(path,label,inicio,oficial):
    wb=openpyxl.load_workbook(path,data_only=True); ws=wb.active
    ef=cr=dep=yape=tarj=egr=0.0
    otros=[]; salidas=[]; mixed=[]; vsum={}
    for r in range(3,ws.max_row+1):
        det=ws.cell(r,2).value
        if det is None: continue
        e=ws.cell(r,4).value or 0; c=ws.cell(r,5).value or 0; d=ws.cell(r,6).value or 0
        chq=ws.cell(r,7).value or 0; t=ws.cell(r,8).value or 0; y=ws.cell(r,10).value or 0
        eg=ws.cell(r,11).value or 0; tipo=ws.cell(r,15).value; orden=ws.cell(r,22).value
        nro=ws.cell(r,3).value
        if orden==1 and tipo=='I':  # ventas
            ef+=e;cr+=c;dep+=d;yape+=y;tarj+=t
            nz=[x for x in [('EF',e),('CR',c),('DEP',d),('YAPE',y),('TARJ',t)] if x[1]]
            if len(nz)>1: mixed.append((nro,det,e,c,d,y,t))
            vsum.setdefault(nro,[]).append((e,c,d,y,t))
        elif orden==2:  # otros ingresos (cobros/vueltos) cash
            otros.append((det,e))
        elif orden==3 or (tipo=='S' and orden==1):  # salidas/egresos
            salidas.append((det,eg,'GI' if orden==1 else 'SAL'))
        egr+= eg if (tipo=='S') else 0
    print(f"=== {label} ===")
    print(f"ventas EFECTIVO={round(ef,2)} CREDITO={round(cr,2)} DEPOSITO={round(dep,2)} YAPE={round(yape,2)} TARJETA={round(tarj,2)} DEP+YAPE={round(dep+yape,2)}")
    print("OTROS INGRESOS(orden2):",otros," sum_ef=",round(sum(o[1] for o in otros),2))
    print("SALIDAS/EGRESOS:",salidas," sum=",round(sum(s[1] for s in salidas),2))
    print("MIXED ventas:",mixed)
    otros_ef=sum(o[1] for o in otros); sal=sum(s[1] for s in salidas)
    fin=inicio+ef+otros_ef-sal
    print(f"fin_calc(inicio {inicio}+ef {round(ef,2)}+otros {round(otros_ef,2)}-salidas {round(sal,2)})={round(fin,2)} oficial={oficial} diff={round(fin-oficial,2)}")
    print()
    return dep,yape,tarj

analyze_system("produccion-09-07/CAJAAA.xlsx","DAY09 SYSTEM (CAJAAA)",11038.79,13591.69)
analyze_system("produccion-10-07/CAJA 10-07.xlsx","DAY10 SYSTEM",13591.69,23903.49)

# ---------- BCP ----------
wb=openpyxl.load_workbook("produccion-10-07/Corriente Soles - 2026-07-10T191711.766.xlsx",data_only=True)
ws=wb['data']; tot=0
print("=== BCP CORRIENTE SOLES ===")
for r in range(2,ws.max_row+1):
    f=ws.cell(r,1).value; d=ws.cell(r,2).value; m=ws.cell(r,4).value; op=ws.cell(r,5).value
    print(f"{f} | {d} | {m} | op{op}"); tot+=m
print("NET movimientos=",round(tot,2))
