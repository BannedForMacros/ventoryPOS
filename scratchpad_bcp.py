import openpyxl
for path in ["produccion-10-07/Corriente Soles - 2026-07-10T191711.766.xlsx","produccion-09-07/CAJAAA.xlsx"]:
    wb = openpyxl.load_workbook(path, data_only=True)
    for sh in wb.sheetnames:
        ws=wb[sh]
        print("=====",path,"| sheet:",sh, ws.max_row,"x",ws.max_column)
        for r in range(1, ws.max_row+1):
            row=[]
            for c in range(1, ws.max_column+1):
                v=ws.cell(r,c).value
                if v is not None: row.append(f"{c}:{v!r}")
            if row: print(f"R{r}", " | ".join(row))
