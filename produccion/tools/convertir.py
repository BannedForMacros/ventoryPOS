#!/usr/bin/env python3
"""
Convierte los .xlsx de produccion/ a JSON limpio y determinista en produccion/data/,
para que el seeder PHP no dependa de un parser de Excel.

Uso:  python3 produccion/tools/convertir.py
Salida: produccion/data/{stock,por_cobrar,por_pagar,anticipos}.json  (+ valida totales)

El balance diario (01-08 jul) se mantiene aparte en produccion/data/balance_diario.json
(escrito a mano desde la investigacion; el .xlsx de 13MB es inestable de leer en bulk).
"""
import openpyxl, datetime, json, os, re, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # produccion/
DATA = os.path.join(BASE, "data")
os.makedirs(DATA, exist_ok=True)


def s(v):
    """string normalizada: sin espacios de relleno, colapsa espacios internos."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    try:
        return round(float(str(v).replace(",", "")), 6)
    except ValueError:
        return 0.0


def fecha(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def norm_prod(v):
    """clave de match de producto: mayusculas, sin espacios de relleno."""
    return re.sub(r"\s+", " ", str(v or "")).strip().upper()


def load(fname, sheet=None):
    wb = openpyxl.load_workbook(os.path.join(BASE, fname), data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def write(name, obj):
    path = os.path.join(DATA, name + ".json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    print(f"  -> data/{name}.json")
    return path


# ---------------------------------------------------------------- STOCK
def do_stock():
    rows = load("STOCK 08-07.xlsx")
    # header en fila 2 (idx1): CODIGO CODI PRODUCTO UNID TIENDA ALMACEN1 TOTAL COSTO VALOR "C U IGV"
    out, seen = [], {}
    for r in rows[2:]:
        prod = s(r[2])
        if not prod:
            continue
        codi = s(r[1])
        cantidad = num(r[6])   # TOTAL
        costo = num(r[7])      # COSTO unitario
        pv = num(r[9])         # C U IGV = precio venta con IGV
        key = codi or prod
        if key in seen:        # dedupe defensivo por codigo
            continue
        seen[key] = True
        out.append({
            "codi": codi,
            "producto": prod,
            "unidad": s(r[3]) or "NIU",
            "cantidad": round(cantidad, 4),
            "costo": round(costo, 4),
            "precio_venta_igv": round(pv, 2),
        })
    total_val = round(sum(x["cantidad"] * x["costo"] for x in out), 2)
    write("stock", out)
    print(f"  STOCK: {len(out)} productos · valorizado a costo = {total_val:,.2f} (Excel balance 08-07: 258,931.52)")
    return out


# ---------------------------------------------------------------- POR COBRAR
def do_por_cobrar():
    rows = load("POR COBRAR 08-07.xlsx")
    # title fila1, header fila2: NRUC CLIENTE DCTO FECHA SALDOSOLES SALDODOLARES
    out = []
    for r in rows[2:]:
        cliente = s(r[1])
        if not cliente or cliente.upper().startswith("TOTAL"):
            continue
        ss = num(r[4]); sd = num(r[5])
        if abs(ss) < 0.005 and abs(sd) < 0.005:
            continue
        out.append({
            "ruc": s(r[0]),
            "cliente": cliente,
            "comprobante": s(r[2]),
            "fecha": fecha(r[3]),
            "saldo_soles": round(ss, 2),
            "saldo_dolares": round(sd, 2),
        })
    tsol = round(sum(x["saldo_soles"] for x in out), 2)
    tdol = round(sum(x["saldo_dolares"] for x in out), 2)
    write("por_cobrar", out)
    print(f"  POR COBRAR: {len(out)} filas · S/ {tsol:,.2f} · US$ {tdol:,.2f} (Excel balance 08-07: 81,866.63)")
    return out


# ---------------------------------------------------------------- PROVEEDORES POR PAGAR
def do_por_pagar():
    rows = load("PROVEEDORES POR PAGAR.xlsx")
    out = []
    for r in rows[2:]:
        prov = s(r[1])
        if not prov or prov.upper().startswith("TOTAL"):
            continue
        ss = num(r[4]); sd = num(r[5])
        out.append({
            "ruc": s(r[0]),
            "proveedor": prov,
            "comprobante": s(r[2]),
            "fecha": fecha(r[3]),
            "saldo_soles": round(ss, 2),
            "saldo_dolares": round(sd, 2),
        })
    tsol = round(sum(x["saldo_soles"] for x in out), 2)
    write("por_pagar", out)
    print(f"  POR PAGAR: {len(out)} filas · S/ {tsol:,.2f} (neg = creditos del proveedor)")
    return out


# ---------------------------------------------------------------- VENTAS (indice de precio real)
def ventas_precio_index():
    """{ (comprobante, producto_norm): precio_unitario_real } desde VENTAS.xlsx (con IGV)."""
    rows = load("VENTAS.xlsx")
    # header fila5 (idx4): NDOC TDOC FECHA CLIENTE CODA CODIGOF PRODUCTO CANT PRECIO IMPORTE ...
    idx = {}
    for r in rows[5:]:
        ndoc = s(r[0]); prod = norm_prod(r[6]); precio = num(r[8])
        if ndoc and prod and precio > 0:
            idx.setdefault((ndoc, prod), precio)
    return idx


# Precios reales confirmados por la ENTREGA del 09-07 (no están en ningún Excel de precios):
# Jibaja entregó el ladrillo P00100036534 valorado 660 → 0.66/u. Fuente: bajada exacta del
# anticipo 08→09 (1,205.50 = 47.50 + 498 + 660) contra el balance diario.
PRECIOS_CONFIRMADOS = {
    ("P00100036534", "LADRILLO ESTANDART 18H FORTES"): 0.66,
}


# ---------------------------------------------------------------- ANTICIPOS (pendientes por entregar)
def do_anticipos(stock):
    # fallback: costo + IGV (columna "C U IGV" del stock) por nombre normalizado
    costo_igv = {}
    for x in stock:
        costo_igv[norm_prod(x["producto"])] = x["precio_venta_igv"]
    ventas = ventas_precio_index()

    rows = load("PENDIENTES POR ENTREGAR 08-07.xlsx")
    grupos = {}   # (cliente, comprobante) -> {..., monto, lineas, exacto}
    n_real = n_costo = n_sin = 0
    for r in rows[5:]:
        cliente = s(r[0]); denom = s(r[1])
        if not cliente or not denom:
            continue
        saldo = num(r[5])              # cantidad pendiente por entregar
        if saldo <= 0:
            continue
        comprobante = s(r[6]); femis = fecha(r[7]); pk = norm_prod(denom)
        # jerarquia: precio confirmado por entrega -> venta real -> costo+IGV -> 0
        pu = PRECIOS_CONFIRMADOS.get((comprobante, pk))
        if pu is None:
            pu = ventas.get((comprobante, pk))
        linea_exacta = pu is not None      # precio real/confirmado (no escalable)
        if pu is None:
            pu = costo_igv.get(pk)
            if pu is None:
                n_sin += 1; pu = 0.0
            else:
                n_costo += 1
        else:
            n_real += 1
        monto = round(saldo * pu, 2)
        key = (cliente, comprobante)
        g = grupos.setdefault(key, {"cliente": cliente, "comprobante": comprobante,
                                    "fecha": femis, "monto": 0.0, "lineas": 0, "exacto": True})
        g["monto"] = round(g["monto"] + monto, 2)
        g["lineas"] += 1
        if not linea_exacta:
            g["exacto"] = False
    out = sorted(grupos.values(), key=lambda g: (g["cliente"], g["comprobante"]))
    total = round(sum(g["monto"] for g in out), 2)
    exacto_sum = round(sum(g["monto"] for g in out if g["exacto"]), 2)
    write("anticipos", out)
    print(f"  ANTICIPOS: {len(out)} comprobantes · valorizado S/ {total:,.2f} "
          f"(Excel balance 08-07 CLIENTES ANTICIPOS: 53,088.74)")
    print(f"    lineas: {n_real} precio real · {n_costo} costo+IGV · {n_sin} sin precio")
    print(f"    exactos (precio real, NO se escalan): {sum(1 for g in out if g['exacto'])} · S/ {exacto_sum:,.2f}")
    return out


if __name__ == "__main__":
    print("Convirtiendo produccion/*.xlsx -> produccion/data/*.json")
    stock = do_stock()
    do_por_cobrar()
    do_por_pagar()
    do_anticipos(stock)
    print("Listo.")
