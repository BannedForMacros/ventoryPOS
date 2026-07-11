#!/usr/bin/env python3
"""
Convierte los .xlsx de produccion-10-07/ (corte 2026-07-10) a JSON en
produccion-10-07/data/. Mismo espíritu que convertir.py (día 08), adaptado a los
nombres/estructura de los reportes nuevos y a la valorización del balance del 10:
  - STOCK valorizado a COSTO + IGV (así lo lleva el balance del 10: 244,461.00).
  - ANTICIPOS: precio de venta real (VENTAS.xlsx del 08, sigue válido por producto)
    donde exista, si no costo+IGV; se marca `exacto` y el seeder escala solo lo estimado
    para cerrar en 52,303.24 (CLIENTES ANTICIPOS del balance 10).
Uso:  python3 produccion/tools/convertir_10.py
"""
import openpyxl, datetime, json, os, re

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))            # produccion/
SRC  = os.path.join(BASE, "..", "produccion-10-07")                          # carpeta día 10
DATA = os.path.join(SRC, "data")
os.makedirs(DATA, exist_ok=True)

# targets del balance diario del 10-07 (para validar)
T_STOCK = 244461.00
T_CXC   = 83015.38
T_CXP   = 136280.98
T_ANTIC = 52303.24


def s(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()

def num(v):
    if v is None or v == "":
        return 0.0
    try:
        return round(float(str(v).replace(",", "")), 6)
    except ValueError:
        return 0.0

def fecha(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.date() if isinstance(v, datetime.datetime) else v).isoformat()
    return None

def norm_prod(v):
    return re.sub(r"\s+", " ", str(v or "")).strip().upper()

def load(path, sheet=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows

def write(name, obj):
    p = os.path.join(DATA, name + ".json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=1)
    return p


# --- STOCK (costo + IGV) --------------------------------------------------
def do_stock():
    rows = load(os.path.join(SRC, "STOCK INVENTARIO.xlsx"))
    # header fila1: CODIGO CODI PRODUCTO UNID TIENDA ALMACEN1 TOTAL COSTO VALOR
    out, seen = [], set()
    for r in rows[2:]:
        prod = s(r[2])
        if not prod:
            continue
        codi = s(r[1]); key = codi or prod
        if key in seen:
            continue
        seen.add(key)
        cant = num(r[6]); costo = num(r[7])
        out.append({
            "codi": codi, "producto": prod, "unidad": s(r[3]) or "NIU",
            "cantidad": round(cant, 4), "costo": round(costo, 4),
            "costo_igv": round(costo * 1.18, 4),
        })
    val_costo = round(sum(x["cantidad"] * x["costo"] for x in out), 2)
    val_igv   = round(sum(x["cantidad"] * x["costo_igv"] for x in out), 2)
    write("stock", out)
    print(f"  STOCK: {len(out)} SKUs · costo {val_costo:,.2f} · costo+IGV {val_igv:,.2f} "
          f"(target {T_STOCK:,.2f} · dif {val_igv - T_STOCK:+.2f} → partida de ajuste)")
    return out


# --- CUENTAS POR COBRAR (agregado por cliente) ----------------------------
def do_por_cobrar():
    rows = load(os.path.join(SRC, "CUENTAS POR COBRAR 1.xlsx"))
    # header fila1: NRUC PROVEEDOR SOLES DOLARES
    out = []
    for r in rows[2:]:
        cli = s(r[1])
        if not cli or cli.upper().startswith("TOTAL"):
            continue
        ss = num(r[2]); sd = num(r[3])
        if abs(ss) < 0.005 and abs(sd) < 0.005:
            continue
        out.append({"ruc": s(r[0]), "cliente": cli, "comprobante": "",
                    "fecha": None, "saldo_soles": round(ss, 2), "saldo_dolares": round(sd, 2)})
    tot = round(sum(x["saldo_soles"] for x in out), 2)
    write("por_cobrar", out)
    print(f"  CxC: {len(out)} clientes · S/ {tot:,.2f} (target {T_CXC:,.2f} · dif {tot - T_CXC:+.2f})")
    return out


# --- CUENTAS POR PAGAR ----------------------------------------------------
def do_por_pagar():
    rows = load(os.path.join(SRC, "CUENTAS POR PAGAR 1.xlsx"))
    # header fila1: NRUC CLIENTE DCTO FECHA SALDOSOLES SALDODOLARES
    out = []
    for r in rows[2:]:
        prov = s(r[1])
        if not prov or prov.upper().startswith("TOTAL"):
            continue
        out.append({"ruc": s(r[0]), "proveedor": prov, "comprobante": s(r[2]),
                    "fecha": fecha(r[3]), "saldo_soles": round(num(r[4]), 2),
                    "saldo_dolares": round(num(r[5]), 2)})
    tot = round(sum(x["saldo_soles"] for x in out), 2)
    write("por_pagar", out)
    print(f"  CxP: {len(out)} filas · S/ {tot:,.2f} (target {T_CXP:,.2f} · dif {tot - T_CXP:+.2f})")
    return out


# --- ANTICIPOS (pendientes por entregar del 10) ---------------------------
def ventas_index():
    rows = load(os.path.join(BASE, "VENTAS.xlsx"))   # ventas del 08 siguen dando precio por producto
    idx = {}
    for r in rows[5:]:
        nd = s(r[0]); pr = norm_prod(r[6]); pv = num(r[8])
        if nd and pr and pv > 0:
            idx.setdefault((nd, pr), pv)
    return idx


def do_anticipos(stock):
    costo_igv = {norm_prod(x["producto"]): x["costo_igv"] for x in stock}
    ventas = ventas_index()
    rows = load(os.path.join(SRC, "PENDIENTE POR ENTREGAR 1.xlsx"))
    # header fila4: Cliente Denominacion Unidad Pedido Entregado Saldo Referencia FechaEmision
    grupos = {}
    n_real = n_costo = n_sin = 0
    for r in rows[5:]:
        cliente = s(r[0]); denom = s(r[1])
        if not cliente or not denom:
            continue
        saldo = num(r[5])
        if saldo <= 0:
            continue
        comp = s(r[6]); femis = fecha(r[7]); pk = norm_prod(denom)
        pu = ventas.get((comp, pk))
        exacta = pu is not None
        if pu is None:
            pu = costo_igv.get(pk)
            if pu is None:
                n_sin += 1; pu = 0.0
            else:
                n_costo += 1
        else:
            n_real += 1
        monto = round(saldo * pu, 2)
        g = grupos.setdefault((cliente, comp), {"cliente": cliente, "comprobante": comp,
                              "fecha": femis, "monto": 0.0, "lineas": 0, "exacto": True})
        g["monto"] = round(g["monto"] + monto, 2); g["lineas"] += 1
        if not exacta:
            g["exacto"] = False
    out = sorted(grupos.values(), key=lambda g: (g["cliente"], g["comprobante"]))
    tot = round(sum(g["monto"] for g in out), 2)
    ex = round(sum(g["monto"] for g in out if g["exacto"]), 2)
    write("anticipos", out)
    print(f"  ANTICIPOS: {len(out)} comprobantes · valorizado {tot:,.2f} (target {T_ANTIC:,.2f})")
    print(f"    {n_real} real · {n_costo} costo+IGV · {n_sin} sin precio · exactos S/ {ex:,.2f}")
    return out


if __name__ == "__main__":
    print("Convirtiendo produccion-10-07/*.xlsx -> produccion-10-07/data/*.json")
    stock = do_stock()
    do_por_cobrar()
    do_por_pagar()
    do_anticipos(stock)
    print("Listo.")
